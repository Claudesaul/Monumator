"""
Daily Stockout Excel Processing
==============================

Handles Excel generation for Daily Stockout Reports using openpyxl.
Preserves template formatting while inserting data.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .base_excel import ExcelProcessorBase

class StockoutExcelProcessor(ExcelProcessorBase):
    """
    Processes Daily Stockout Report Excel files using openpyxl
    """
    
    def __init__(self, template_path="templates/Daily Stockout Report.xlsx"):
        """Initialize with Daily Stockout template"""
        super().__init__(template_path)
        self.workbook = None
        
    def load_workbook(self):
        """
        Load the working copy workbook
        
        Returns:
            openpyxl.Workbook: Loaded workbook object
        """
        if not self.output_path:
            raise FileNotFoundError("Working copy not found. Call create_working_copy() first.")
        
        try:
            self.workbook = load_workbook(self.output_path)
            return self.workbook
        except Exception as e:
            print(f"❌ Failed to load workbook: {str(e)}")
            raise
    
    def insert_dataframe_to_sheet(self, dataframe, sheet_name, start_row=2, start_col=1):
        """
        Insert a pandas DataFrame into a worksheet
        
        Args:
            dataframe (pandas.DataFrame): Data to insert
            sheet_name (str): Name of the target sheet
            start_row (int): Starting row for data insertion
            start_col (int): Starting column for data insertion
        """
        if not self.workbook:
            raise RuntimeError("Workbook not loaded. Call load_workbook() first.")
        
        if sheet_name not in self.workbook.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found in workbook")
            return
        
        if dataframe.empty:
            return
        
        sheet = self.workbook[sheet_name]
        
        # Convert DataFrame to rows (without header)
        for row_idx, row_data in enumerate(dataframe_to_rows(dataframe, index=False, header=False), start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                sheet.cell(row=row_idx, column=col_idx, value=value)
    
    def populate_sheets(self, data_dict):
        """
        Populate all sheets with data
        
        Args:
            data_dict (dict): Dictionary containing data for each sheet
        """
        sheet_mappings = {
            'highlights': 'Highlights',
            'markets': 'Markets', 
            'null_orders': 'NullOrders',
            'ocs': 'OCS'
        }
        
        for data_key, sheet_name in sheet_mappings.items():
            if data_key in data_dict:
                data = data_dict[data_key]
                # Insert new data directly (template is always blank)
                self.insert_dataframe_to_sheet(data, sheet_name)
    
    def save_workbook(self):
        """
        Save the workbook and close it
        
        Returns:
            str: Path to the saved file
        """
        if not self.workbook:
            raise RuntimeError("Workbook not loaded")
        
        try:
            self.workbook.save(self.output_path)
            self.workbook.close()
            return self.output_path
        except Exception as e:
            print(f"❌ Failed to save workbook: {str(e)}")
            raise
    
    def generate_stockout_report(self, data_dict, output_directory="downloads/daily"):
        """
        Complete workflow to generate Daily Stockout Report
        
        Args:
            data_dict (dict): Dictionary containing all sheet data
            output_directory (str): Directory to save the report
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            # Create working copy
            self.create_working_copy(output_directory, "Daily Stockout Report")
            
            # Load workbook
            self.load_workbook()
            
            # Populate all sheets
            self.populate_sheets(data_dict)
            
            # Save and return path
            return self.save_workbook()
            
        except Exception as e:
            print(f"❌ Failed to generate stockout report: {str(e)}")
            raise